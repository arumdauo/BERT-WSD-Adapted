
import openpyxl
import re

""" 
    Gets information about lemmas in the AAT dataset
    - AAT_dataset_path:           path of the AAT dataset .xlsx file

"""

def _get_info(lemma, AAT_dataset_path):

    db = openpyxl.load_workbook(AAT_dataset_path)
    sheet_name = db.sheetnames[0]
    ws = db[sheet_name]
    global aat_id, aat_lemma, aat_lemma_eccezioni, aat_glossa
    
    results = {}
   
    for i in range(2, ws.max_row+1):
        if ws.cell(row=i, column=3).value != None:
            lemma_names = get_all_lemma_names(ws.cell(row=i, column=3).value)
            lemma_names_2 = get_all_lemma_names(ws.cell(row=i, column=7).value)
        if ws.cell(row=i, column=4).value != None:
            exceptions = ws.cell(row=i, column=4).value
            exceptions_2 = ws.cell(row=i, column=8).value
            lemma_exc_names = get_all_lemma_names(exceptions)
            lemma_exc_names_2 =  get_all_lemma_names(exceptions_2)
            if (any(lemma == l for l in lemma_exc_names) or 
                any (lemma == l for l in lemma_exc_names_2)):
                term_id = ws.cell(row=i, column=1).value
                term = ws.cell(row=i, column=4).value
                term = re.sub(r'\([^)]*\)', '', term)
                term = term.rstrip()
                term = lemma + "%" + term_id
                results[term] = ws.cell(row=i, column=6).value
        if (ws.cell(row=i, column=3).value == lemma or 
            any(lemma == l for l in lemma_names) or 
            any(lemma == l for l in lemma_names_2)):
            term_id = ws.cell(row=i, column=1).value
            term = ws.cell(row=i, column=3).value
            term = re.sub(r'\([^)]*\)', '', term)
            term = term.rstrip()
            term = lemma + "%" + term_id
            results[term] = ws.cell(row=i, column=6).value          
    return results
            
def get_glosses(lemma, AAT_dataset_path):
    return _get_info(lemma, AAT_dataset_path)

def get_example_sentences(lemma, pos):
    return _get_info(lemma, pos, info_type='examples')

def get_all_lemma_names(stringa):
    regex_par = '[\(\[].*?[\)\]]'
    stringa = re.sub(regex_par, "", stringa)
    regex = r'\w+'
    lemma_names = re.findall(regex.lower(), stringa.lower())
    return lemma_names

    